import json, os
from collections import defaultdict
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

OUT = "/Users/ln123/Document/pythonfile/memory"

# ---- Paper Table 4 data (15 models) ----
# Columns: KE, KA, KH, PB, PU, CC_Avg, BT | Flu, Coh, Cons, CA_Avg | HL, CS, ED, Emp, RA_Avg
paper = {
    "ChatGLM3-6B":   [2.016,2.792,2.704,2.455,2.812, 0.532, 3.269,3.647,3.283, 3.064,2.932,1.969,2.993],
    "XVERSE-7B":     [1.834,2.774,2.763,2.564,2.887, 0.620, 3.393,3.752,3.518, 3.395,2.743,2.013,2.936],
    "Baichuan2-7B":  [1.813,2.849,2.929,2.830,3.081, 0.625, 3.551,3.894,3.827, 3.670,2.728,2.115,2.984],
    "Qwen-7B":       [1.956,2.728,2.633,2.605,2.780, 0.606, 3.187,3.564,3.229, 3.036,2.791,2.052,2.838],
    "InternLM-7B":   [1.782,2.800,2.781,2.719,3.016, 0.630, 3.527,3.823,3.744, 3.546,2.622,2.070,2.897],
    "XVERSE-13B":    [1.977,2.828,2.862,2.579,2.915, 0.630, 3.444,3.811,3.559, 3.319,2.939,2.045,3.018],
    "Baichuan2-13B": [1.802,2.869,2.946,2.808,3.081, 0.639, 3.596,3.924,3.864, 3.700,2.703,2.136,3.021],
    "Qwen-14B":      [1.988,2.800,2.811,2.744,2.900, 0.620, 3.351,3.765,3.510, 3.354,2.871,2.237,2.970],
    "InternLM-20B":  [1.945,2.916,2.920,2.753,3.041, 0.648, 3.576,3.943,3.717, 3.582,2.885,2.132,3.047],
    "CharacterGLM":  [1.640,2.819,2.738,2.301,2.969, None,  3.414,3.717,3.737, 3.738,2.265,1.966,2.812],
    "Xingchen":      [1.636,2.768,2.743,2.772,3.055, 0.630, 3.378,3.807,3.754, 3.757,2.272,2.100,2.799],
    "MiniMax":       [1.835,2.910,2.944,2.774,3.125, 0.685, 3.609,3.932,3.811, 3.768,2.672,2.150,3.017],
    "BC-NPC-Turbo":  [1.802,2.964,2.993,2.910,3.151, 0.681, 3.578,3.898,3.916, 3.836,2.643,2.336,2.971],
    "GPT-3.5":       [1.716,2.339,2.212,1.921,2.316, 0.653, 2.629,2.917,2.700, 2.565,2.422,1.660,2.526],
    "GPT-4":         [2.250,2.855,2.785,2.721,2.873, 0.694, 3.332,3.669,3.343, 3.143,3.184,2.153,3.010],
}

# ---- Our system data from evaluation_scores.xlsx ----
wb_in = openpyxl.load_workbook("/Users/ln123/Desktop/evaluation_scores.xlsx", data_only=True)
ws_sum = wb_in["汇总"]
# bottom rows: metric averages and metric names
metric_avg = None
metric_names = None
for r in ws_sum.iter_rows(min_row=3, values_only=True):
    vals = [v for v in r if v is not None]
    if len(vals)==12 and all(isinstance(v,(int,float)) for v in vals):
        metric_avg = vals
    if len(vals)==12 and all(isinstance(v,str) for v in vals):
        metric_names = vals

# 汇总 column order: KA,CS,Coh,ED,Flu,KE,PU,Cons,PB,KH,Emp,HL
our_raw = dict(zip(metric_names, metric_avg))
# map to paper keys
our = {
    "KE":  our_raw["知识曝光度"], "KA": our_raw["知识准确率"], "KH": our_raw["知识幻觉性"],
    "PB":  our_raw["行为一致性"], "PU": our_raw["言语一致性"],
    "Flu": our_raw["对话流利度"], "Coh": our_raw["对话连贯性"], "Cons": our_raw["对话一致性"],
    "HL":  our_raw["类人程度"], "CS": our_raw["交流技巧"], "ED": our_raw["表现多样性"], "Emp": our_raw["共情度"],
}
our_cc_avg = round(mean([our["KE"],our["KA"],our["KH"],our["PB"],our["PU"]]), 3)
our_ca_avg = round(mean([our["Flu"],our["Coh"],our["Cons"]]), 3)
our_ra_avg = round(mean([our["HL"],our["CS"],our["ED"],our["Emp"]]), 3)
our_overall = round(mean([our_cc_avg, our_ca_avg, our_ra_avg]), 3)
print(f"Our system: CC={our_cc_avg}, CA={our_ca_avg}, RA={our_ra_avg}, Overall={our_overall}")
print(f"  KE={our['KE']:.3f} KA={our['KA']:.3f} KH={our['KH']:.3f} PB={our['PB']:.3f} PU={our['PU']:.3f}")
print(f"  Flu={our['Flu']:.3f} Coh={our['Coh']:.3f} Cons={our['Cons']:.3f}")
print(f"  HL={our['HL']:.3f} CS={our['CS']:.3f} ED={our['ED']:.3f} Emp={our['Emp']:.3f}")

# compute dimension averages for paper models too
for m, v in paper.items():
    cc = round(mean(v[0:5]), 3)
    ca = round(mean(v[6:9]), 3)
    ra = round(mean(v[9:13]), 3)
    ov = round(mean([cc,ca,ra]), 3)
    v.extend([cc, ca, ra, ov])

# ---- Build Excel ----
wb = Workbook()
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF", size=9)
sub_fill = PatternFill("solid", fgColor="8EAADB")
sub_font = Font(bold=True, color="FFFFFF", size=9)
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
ours_fill = PatternFill("solid", fgColor="FFF2CC")
wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

# model order: ours first, then paper models
ALL_MODELS = ["Ours"] + list(paper.keys())

ws = wb.active
ws.title = "Table4-Comparison"

# Title row
ws.merge_cells("A1:R1")
ws["A1"] = "CharacterEval 实验对比：本系统 vs 论文 Table 4（满分 5，越高越好）"
ws["A1"].font = Font(bold=True, size=12)

# Group header row (row 3)
groups = [
    ("模型", 1, 1),
    ("Character Consistency 角色一致性", 2, 8),
    ("Conversational Ability 对话能力", 9, 13),
    ("Role-playing Attractiveness 角色扮演吸引力", 14, 18),
]
for label, c1, c2 in groups:
    ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
    cell = ws.cell(3, c1, label)
    cell.fill = sub_fill; cell.font = sub_font; cell.alignment = center; cell.border = border

# Metric header row (row 4)
headers = ["模型", "KE", "KA", "KH", "PB", "PU", "Avg.", "BT",
           "Flu.", "Coh.", "Cons.", "Avg.",
           "HL", "CS", "ED", "Emp.", "Avg.", "Overall"]
for j, h in enumerate(headers, start=1):
    cell = ws.cell(4, j, h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

# Metric full-name row (row 5)
full_names = ["", "知识曝光", "知识准确", "知识幻觉", "行为一致", "言语一致", "维度均", "人格回测",
              "流利度", "连贯性", "对话一致", "维度均",
              "拟人化", "沟通技巧", "表达多样", "共情度", "维度均", "三维度均"]
for j, h in enumerate(full_names, start=1):
    cell = ws.cell(5, j, h)
    cell.font = Font(size=8, italic=True); cell.alignment = center; cell.border = border

# Data rows
r = 6
for model in ALL_MODELS:
    if model == "Ours":
        vals = [our["KE"],our["KA"],our["KH"],our["PB"],our["PU"], our_cc_avg, None,
                our["Flu"],our["Coh"],our["Cons"], our_ca_avg,
                our["HL"],our["CS"],our["ED"],our["Emp"], our_ra_avg, our_overall]
    else:
        v = paper[model]
        vals = [v[0],v[1],v[2],v[3],v[4], v[13], v[5], v[6],v[7],v[8], v[14], v[9],v[10],v[11],v[12], v[15], v[16]]
    ws.cell(r, 1, model)
    for j, val in enumerate(vals, start=2):
        cell = ws.cell(r, j)
        if val is None:
            cell.value = "-"
        else:
            cell.value = round(val, 3)
        cell.alignment = center
        cell.border = border
    ws.cell(r, 1).border = border
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    if model == "Ours":
        for j in range(1, len(headers)+1):
            ws.cell(r, j).fill = ours_fill
            ws.cell(r, j).font = bold
    r += 1

# bold best per metric column
for j in range(2, len(headers)+1):
    col_vals = []
    for rr in range(6, 6+len(ALL_MODELS)):
        v = ws.cell(rr, j).value
        if isinstance(v, (int, float)):
            col_vals.append((rr, v))
    if col_vals:
        mx = max(v for _, v in col_vals)
        for rr, v in col_vals:
            if abs(v - mx) < 0.0005:
                ws.cell(rr, j).font = Font(bold=True)

# notes
r += 1
notes = [
    "说明：",
    "1. 数据来源：CharacterEval 论文 Table 4（15 个模型）+ 本系统评测结果（evaluation_scores.xlsx，77 个角色）。",
    "2. 论文将 13 个指标归为 3 个可见维度 + 人格回测。本系统缺人格回测（BT），以 \"-\" 表示。",
    "3. KE=知识曝光度, KA=知识准确率, KH=知识幻觉性, PB=行为一致性, PU=言语一致性。",
    "4. Flu=流利度, Coh=连贯性, Cons=对话一致性, HL=拟人化, CS=沟通技巧, ED=表达多样性, Emp=共情度。",
    "5. Avg.=该维度指标均值; Overall=三个维度均值的再平均。加粗为该列最高分。",
]
for note in notes:
    ws.cell(r, 1, note).font = Font(size=9)
    r += 1

# column widths
ws.column_dimensions["A"].width = 16
for j in range(2, len(headers)+1):
    ws.column_dimensions[get_column_letter(j)].width = 7

ws.freeze_panes = "B6"

# ---- Sheet 2: dimension summary ----
ws2 = wb.create_sheet("Dimension-Summary")
ws2.merge_cells("A1:E1")
ws2["A1"] = "维度平均分对比"
ws2["A1"].font = Font(bold=True, size=12)
dim_headers = ["模型", "角色一致性", "对话能力", "角色扮演吸引力", "Overall"]
for j, h in enumerate(dim_headers, start=1):
    cell = ws2.cell(2, j, h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

r = 3
# sort by overall descending
model_dims = [("Ours", our_cc_avg, our_ca_avg, our_ra_avg, our_overall)]
for m, v in paper.items():
    model_dims.append((m, v[13], v[14], v[15], v[16]))
model_dims.sort(key=lambda x: -x[4])

for m, cc, ca, ra, ov in model_dims:
    ws2.cell(r, 1, m)
    ws2.cell(r, 2, round(cc, 3))
    ws2.cell(r, 3, round(ca, 3))
    ws2.cell(r, 4, round(ra, 3))
    ws2.cell(r, 5, round(ov, 3))
    for j in range(1, 6):
        ws2.cell(r, j).border = border
        ws2.cell(r, j).alignment = center
    if m == "Ours":
        for j in range(1, 6):
            ws2.cell(r, j).fill = ours_fill
            ws2.cell(r, j).font = bold
    r += 1
# bold best
for j in range(2, 6):
    col_vals = [(rr, ws2.cell(rr, j).value) for rr in range(3, 3+len(model_dims))]
    mx = max(v for _, v in col_vals)
    for rr, v in col_vals:
        if abs(v - mx) < 0.0005:
            ws2.cell(rr, j).font = Font(bold=True)

ws2.column_dimensions["A"].width = 16
for col in "BCDE":
    ws2.column_dimensions[col].width = 18
ws2.freeze_panes = "A3"

xlsx = os.path.join(OUT, "Table4_comparison.xlsx")
wb.save(xlsx)
print("Saved:", xlsx)
