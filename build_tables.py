import json, os
from collections import defaultdict
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/ln123/Downloads/compared_results"
OUT = "/Users/ln123/Document/pythonfile/memory"

FILES = {
    "Ours":        "ours/evaluation.jsonl",
    "Qwen3-32B":   "qwen3-32b/evaluation_qwen3_32b_fast.jsonl",
    "InternLM-7B": "internlm-7b/evaluation_internlm_7b.jsonl",
    "Baichuan-7B": "baichuan-7b/evaluation_baichuan_7b.jsonl",
    "ChatGLM3-6B": "chatglm3-6b/evaluation_chatglm3_6b.jsonl",
    "Qwen-7B":     "qwen-7b/evaluation_qwen_7b.jsonl",
    "Xverse-7B":   "xverse-7b/evaluation_xverse_7b.jsonl",
}

METRICS = ["Accuracy", "Exposure", "Hallucination",
           "Consistency", "Behavior", "Utterance",
           "Coherence", "Empathy", "Communication_skills",
           "Diversity", "Fluency", "Humanlikeness"]

DIM = {
    "Knowledge": ["Accuracy", "Exposure", "Hallucination"],
    "Persona":   ["Consistency", "Behavior", "Utterance"],
    "Coherence": ["Coherence", "Empathy", "Communication_skills"],
    "Style":     ["Diversity", "Fluency"],
}

# batch 1 + batch 2 = 20 representative characters
PICK = ["佟湘玉", "曾小贤", "孙悟空", "贾宝玉", "杨过",
        "萧炎", "梅长苏", "高启强", "罗辑", "甄嬛",
        "花千骨", "徐凤年", "关宏峰", "武松", "朱朝阳",
        "景天", "李云龙", "顾里", "柯景腾", "侯亮平"]

def load(path):
    d = json.load(open(os.path.join(ROOT, path)))
    by_role = defaultdict(lambda: defaultdict(list))
    novel = {}
    for r in d:
        role = r["role"]; novel[role] = r["novel_name"]
        for m in METRICS:
            if m in r and r[m] is not None:
                by_role[role][m].append(r[m])
    res = {}
    for role, mm in by_role.items():
        res[role] = {m: mean(mm[m]) for m in METRICS if mm[m]}
        res[role]["Overall"] = mean(res[role][m] for m in METRICS if m in res[role])
        for dn, ms in DIM.items():
            res[role][dn] = mean(res[role][m] for m in ms if m in res[role])
    return res, novel

models = {}
novel = {}
for m, f in FILES.items():
    models[m], n = load(f)
    novel.update(n)

roles_all = list(novel.keys())

wb = Workbook()
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF")
bold = Font(bold=True)
center = Alignment(horizontal="center")
ours_fill = PatternFill("solid", fgColor="FFF2CC")

def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

def style_body(ws, r0=2):
    for r in ws.iter_rows(min_row=r0):
        for c in r:
            c.border = border; c.alignment = center

def highlight_ours_col(ws, col_idx, r0, r1):
    for rr in range(r0, r1+1):
        ws.cell(rr, col_idx).fill = ours_fill

def best_in_row(ws, r0, r1, c0, c1):
    # bold the max per data row
    for rr in range(r0, r1+1):
        vals = [(cc, ws.cell(rr, cc).value) for cc in range(c0, c1+1)]
        valid = [(cc, v) for cc, v in vals if isinstance(v, (int, float))]
        if not valid: continue
        mx = max(v for _, v in valid)
        for cc, v in valid:
            if abs(v - mx) < 1e-9:
                ws.cell(rr, cc).font = bold

# ============ Sheet 1: Overall (20 characters) ============
ws = wb.active; ws.title = "Overall"
ws.append(["角色", "作品"] + list(FILES.keys()))
style_header(ws)
for role in PICK:
    ws.append([role, novel[role]] + [round(models[m][role]["Overall"], 2) for m in FILES])
avg_row = ws.max_row + 1
ws.append(["平均", "全部77角色"] + [round(mean(models[m][r]["Overall"] for r in roles_all), 2) for m in FILES])
style_body(ws, 2)
for c in ws[avg_row]:
    c.font = bold
highlight_ours_col(ws, 3, 2, avg_row)
best_in_row(ws, 2, avg_row-1, 3, 3+len(FILES)-1)
ws.append([])
ws.append(["注：综合分为该角色在 12 项 CharacterEval 指标上的平均（满分 5，越高越好）。加粗为该行最高分。"])

# ============ Sheet 2: Dimensions ============
ws2 = wb.create_sheet("Dimensions")
ws2.append(["维度"] + list(FILES.keys()))
style_header(ws2)
dim_pairs = [("Knowledge（知识/幻觉）", "Knowledge"), ("Persona（人格一致）", "Persona"),
             ("Coherence（连贯/共情）", "Coherence"), ("Style（多样/流畅）", "Style"),
             ("Overall（综合）", "Overall")]
for dn, dk in dim_pairs:
    if dk == "Overall":
        vals = [round(mean(models[m][r]["Overall"] for r in roles_all), 2) for m in FILES]
    else:
        vals = [round(mean(models[m][r][dk] for r in roles_all), 2) for m in FILES]
    ws2.append([dn] + vals)
style_body(ws2, 2)
highlight_ours_col(ws2, 2, 2, ws2.max_row)
best_in_row(ws2, 2, ws2.max_row, 2, 1+len(FILES))
ws2.append([])
ws2.append(["注：各维度为该维度下指标在全部 77 角色上的平均分（满分 5）。"])

# ============ Sheet 3: per-metric overall ============
ws3 = wb.create_sheet("Metrics-Overall")
ws3.append(["指标"] + list(FILES.keys()))
style_header(ws3)
for mname in METRICS:
    vals = [round(mean(models[m][r][mname] for r in roles_all if mname in models[m][r]), 2) for m in FILES]
    ws3.append([mname] + vals)
style_body(ws3, 2)
highlight_ours_col(ws3, 2, 2, ws3.max_row)
best_in_row(ws3, 2, ws3.max_row, 2, 1+len(FILES))

# ============ Sheet 4: per character x metric ============
ws4 = wb.create_sheet("PerCharacter")
cur = 1
for role in PICK:
    ws4.cell(cur, 1, f"{role}（{novel[role]}）").font = Font(bold=True, size=12)
    hr = cur + 1
    ws4.cell(hr, 1, "指标")
    for j, m in enumerate(FILES, start=2):
        ws4.cell(hr, j, m)
    style_header(ws4, hr)
    for i, mname in enumerate(METRICS + ["Overall"], start=hr+1):
        ws4.cell(i, 1, mname)
        for j, m in enumerate(FILES, start=2):
            val = models[m][role].get("Overall") if mname == "Overall" else models[m][role].get(mname)
            ws4.cell(i, j, round(val, 2) if val is not None else None)
    cur = hr + len(METRICS) + 2 + 1

# ============ Sheet 5: all roles appendix ============
ws5 = wb.create_sheet("AllRoles")
ws5.append(["角色", "作品"] + list(FILES.keys()))
style_header(ws5)
for role in sorted(roles_all, key=lambda x: -models["Ours"][x]["Overall"]):
    ws5.append([role, novel[role]] + [round(models[m][role]["Overall"], 2) for m in FILES])
style_body(ws5, 2)

for wsx in [ws, ws2, ws3, ws4, ws5]:
    for col in wsx.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        wsx.column_dimensions[get_column_letter(col[0].column)].width = max(ml + 2, 10)
    wsx.freeze_panes = "A2"

xlsx = os.path.join(OUT, "CharacterEval_comparison.xlsx")
wb.save(xlsx)
print("Saved:", xlsx)
